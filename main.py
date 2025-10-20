# main.py
from fastapi import FastAPI, HTTPException
from pydantic import BaseModel
from sqlalchemy.orm import Session
from database import SessionLocal, engine
from models import Base, Response
from typing import List
from datetime import datetime
import sqlite3
from fastapi.responses import FileResponse
import openpyxl
import os
import pandas as pd
from openpyxl.styles import Alignment
from openpyxl.cell.cell import MergedCell


# Crear las tablas si no existen
Base.metadata.create_all(bind=engine)

app = FastAPI(title="NOSACQ-50 API", version="1.0")
@app.get("/")
def home():
    return {"message": "API funcionando correctamente 🚀"}


# --- Secciones de preguntas ---
sections = {
    range(1, 10): "1. Management safety priority and ability",
    range(10, 17): "2. Management safety empowerment",
    range(17, 23): "3. Management safety justice",
    range(23, 29): "4. Workers' safety commitment",
    range(29, 36): "5. Workers safety priority and risk non-acceptance",
    range(36, 44): "6. Safety communication, learning and trust in co-workers",
    range(44, 51): "7. Workers trust in efficiency of safety systems"
}

# --- Modelos de datos ---
class Answer(BaseModel):
    question_number: int
    value: int

class ResponseInput(BaseModel):
    player_id: str
    responses: List[Answer]

# --- Dependencia de sesión ---
def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()

# --- Endpoint principal ---
@app.post("/submit")
def submit_responses(data: ResponseInput):
    db = next(get_db())

    for answer in data.responses:
        section_name = next(
            (name for rng, name in sections.items() if answer.question_number in rng),
            "Unknown Section"
        )

        new_response = Response(
            player_id=data.player_id,
            question_number=answer.question_number,
            value=answer.value,
            section=section_name,
            date=datetime.utcnow()
        )
        db.add(new_response)

    db.commit()
    return {"message": "✅ Respuestas guardadas correctamente", "player": data.player_id}

# --- Endpoint para ver los resultados ---
@app.get("/results")
def get_all_results():
    db = next(get_db())
    results = db.query(Response).all()
    return results

# tabla de resultados en HTML
from fastapi.responses import HTMLResponse

@app.get("/results_table", response_class=HTMLResponse)
def show_results_table():
    conn = sqlite3.connect("responses.db")
    cur = conn.cursor()
    cur.execute("SELECT player_id, question_number, value FROM responses ORDER BY player_id, question_number")
    rows = cur.fetchall()
    conn.close()

    # Agrupar respuestas por jugador
    jugadores = {}
    for player_id, question_number, value in rows:
        if player_id not in jugadores:
            jugadores[player_id] = [None] * 50
        jugadores[player_id][question_number - 1] = value

    # Rango de preguntas por sección
    secciones = [
        ("1. Management safety priority and ability", 1, 9),
        ("2. Management safety empowerment", 10, 16),
        ("3. Management safety justice", 17, 22),
        ("4. Workers' safety commitment", 23, 28),
        ("5. Workers safety priority and risk non-acceptance", 29, 35),
        ("6. Safety communication, learning and trust in co-workers safety competence", 36, 43),
        ("7. Workers trust in efficiency of safety systems", 44, 50),
    ]

    # Modificar el HTML para incluir scripts y estilos
    html = """
    <html>
    <head>
        <title>Resultados NOSACQ-50</title>
        <style>
            body { background:#1e1e2f; color:#eee; font-family:Arial, sans-serif; padding:30px; }
            table { border-collapse: collapse; width:100%; background:#2c2c3e; border-radius:8px; overflow:hidden; }
            th, td { border:1px solid #444; text-align:center; padding:6px; font-size:13px; }
            th { background:#3c3c5c; color:#fff; }
            tr:nth-child(even) { background:#252536; }
            tr:hover { background:#3a3a50; }
            h1 { text-align:center; margin-bottom:20px; }
            .delete-btn { 
                background: #ff4444; 
                color: white; 
                border: none; 
                padding: 5px 10px; 
                border-radius: 4px; 
                cursor: pointer; 
            }
            .delete-btn:hover { 
                background: #ff0000; 
            }
            .export-btn { 
                background: #4CAF50; 
                color: white; 
                border: none; 
                padding: 10px 20px; 
                border-radius: 4px; 
                cursor: pointer;
                margin-bottom: 20px;
            }
            .export-btn:hover { 
                background: #45a049; 
            }
        </style>
        <script>
            async function deletePlayer(playerId) {
                if (confirm('¿Estás seguro de que deseas eliminar este jugador?')) {
                    try {
                        const response = await fetch(`/delete/${playerId}`, {
                            method: 'DELETE'
                        });
                        if (response.ok) {
                            location.reload();
                        } else {
                            alert('Error al eliminar el jugador');
                        }
                    } catch (error) {
                        console.error('Error:', error);
                        alert('Error al eliminar el jugador');
                    }
                }
            }
        </script>
    </head>
    <body>
        <h1>Resultados NOSACQ-50</h1>
        <button class="export-btn" onclick="window.location.href='/export-excel'">📥 Exportar a Excel</button>
        <table>
            <tr>
                <th rowspan="2">Jugador</th>
                <th rowspan="2">Acciones</th>
    """

    # 🔹 Fila de secciones (usa colspan según rango)
    for nombre, start, end in secciones:
        span = end - start + 1
        html += f'<th colspan="{span}">{nombre}</th>'
    html += "</tr>\n<tr>"

    # 🔹 Fila de preguntas (P1–P50)
    for i in range(1, 51):
        html += f"<th>P{i}</th>"
    html += "</tr>"

    # Modificar la parte donde se generan las filas de jugadores
    for player_id, answers in jugadores.items():
        html += f'''<tr>
            <td>{player_id}</td>
            <td><button class="delete-btn" onclick="deletePlayer('{player_id}')">🗑️ Eliminar</button></td>'''
        for val in answers:
            html += f"<td>{val if val is not None else ''}</td>"
        html += "</tr>"

    html += "</table></body></html>"
    return HTMLResponse(content=html)

# Primero agregamos un nuevo endpoint para eliminar registros
@app.delete("/delete/{player_id}")
async def delete_player(player_id: str):
    db = next(get_db())
    db.query(Response).filter(Response.player_id == player_id).delete()
    db.commit()
    return {"message": f"Player {player_id} deleted successfully"}

@app.get("/export-excel")
async def export_excel():
    db = next(get_db())
    results = db.query(Response).all()
    
    # Organizar los datos por jugador
    jugadores = {}
    for result in results:
        if result.player_id not in jugadores:
            jugadores[result.player_id] = {}
        jugadores[result.player_id][f"P{result.question_number}"] = result.value

    # Crear DataFrame con nombres de columnas válidos
    columns = ["Jugador"] + [f"P{i}" for i in range(1, 51)]
    rows = []
    for player_id, answers in jugadores.items():
        row = [player_id]
        for i in range(1, 51):
            row.append(answers.get(f"P{i}", None))
        rows.append(row)
    
    df = pd.DataFrame(rows, columns=columns)

    # Crear Excel
    writer = pd.ExcelWriter('resultados_nosacq50.xlsx', engine='openpyxl')
    df.to_excel(writer, sheet_name='Resultados', index=False, startrow=1)
    
    worksheet = writer.sheets['Resultados']

    # Definir secciones con rangos de columnas válidos
    secciones = [
        ("1. Management safety priority and ability", 'B', 'J'),
        ("2. Management safety empowerment", 'K', 'Q'),
        ("3. Management safety justice", 'R', 'W'),
        ("4. Workers' safety commitment", 'X', 'AD'),
        ("5. Workers safety priority and risk non-acceptance", 'AE', 'AK'),
        ("6. Safety communication, learning and trust in co-workers", 'AL', 'AS'),
        ("7. Workers trust in efficiency of safety systems", 'AT', 'AY')
    ]

    # Agregar encabezados de secciones
    for titulo, inicio, fin in secciones:
        rango = f"{inicio}1:{fin}1"
        worksheet.merge_cells(rango)
        celda = worksheet[f"{inicio}1"]
        celda.value = titulo
        celda.alignment = Alignment(horizontal='center')

    # Establecer un ancho fijo para todas las columnas de preguntas
    FIXED_WIDTH = 8  # Ancho fijo para columnas de preguntas
    PLAYER_COLUMN_WIDTH = 15  # Ancho para la columna de jugador

    # Ajustar la columna del jugador (columna A)
    worksheet.column_dimensions['A'].width = PLAYER_COLUMN_WIDTH

    # Función para generar nombres de columnas de Excel
    def get_column_letter(n):
        result = ''
        while n > 0:
            n, remainder = divmod(n - 1, 26)
            result = chr(65 + remainder) + result
        return result

    # Ajustar todas las columnas de preguntas
    for i in range(1, 51):  # 50 preguntas
        col_letter = get_column_letter(i + 1)  # +1 porque la columna A es para jugador
        worksheet.column_dimensions[col_letter].width = FIXED_WIDTH

    writer.close()

    return FileResponse(
        path="resultados_nosacq50.xlsx",
        filename="resultados_nosacq50.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

